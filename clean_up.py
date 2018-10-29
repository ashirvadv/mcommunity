from openpyxl import Workbook
from openpyxl import load_workbook

def get_old_names():
	'''Get old names.'''
	wb = load_workbook('uniqnames.xlsx')
	sheet = wb['sheet1']
	old_names = set()
	for row in sheet.rows:
		old_names.add(row[0].value)
	return old_names

def get_new_names():
	'''Get new names.'''
	f = open('HELLO.txt', 'r')
	new_names = set()
	for line in f:
		new_names.add(line)
	f.close()
	return new_names

def get_difference(old, new):
	'''Get difference.'''
	return new - old

def write_to_file(diff, filename):
	'''Write to file.'''
	f = open(filename, 'w')
	for name in diff:
		f.write(name)
	f.close()

def update_old_names(new_names):
	'''Update old names.'''
	wb = load_workbook('uniqnames.xlsx')
	sheet = wb['sheet1']
	counter = 1
	for name in new_names:
		key = 'A' + str(counter)
		sheet[key] = name
		counter += 1

def main():
	'''Run main.'''
	old_names = get_old_names()
	new_names = get_new_names()
	difference = get_difference(old_names, new_names)
	write_to_file(difference, 'new.txt')
	update_old_names(new_names)
main()